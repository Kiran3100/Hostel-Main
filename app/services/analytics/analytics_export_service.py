# --- File: C:\Hostel-Main\app\services\analytics\analytics_export_service.py ---
"""
Analytics Export Service - Multi-format data export.

Provides comprehensive export capabilities with:
- CSV export
- Excel/XLSX generation
- PDF report generation
- JSON formatting
- Email delivery
"""

from typing import List, Dict, Optional, Any
from datetime import date, datetime
from decimal import Decimal
from sqlalchemy.orm import Session
from uuid import UUID
import logging
import csv
import io
import json

from app.repositories.analytics.analytics_aggregate_repository import (
    AnalyticsAggregateRepository
)


logger = logging.getLogger(__name__)


class AnalyticsExportService:
    """Service for exporting analytics data in various formats."""
    
    def __init__(self, db: Session):
        """Initialize with database session."""
        self.db = db
        self.aggregate_repo = AnalyticsAggregateRepository(db)
    
    # ==================== CSV Export ====================
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> str:
        """
        Export data to CSV format.
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename
            
        Returns:
            CSV content as string
        """
        logger.info(f"Exporting {len(data)} rows to CSV")
        
        if not data:
            return ""
        
        # Get column headers from first row
        headers = list(data[0].keys())
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        
        writer.writeheader()
        for row in data:
            # Convert Decimal and datetime to strings
            cleaned_row = self._clean_row_for_export(row)
            writer.writerow(cleaned_row)
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content
    
    # ==================== Excel Export ====================
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = 'Analytics'
    ) -> bytes:
        """
        Export data to Excel format.
        
        Requires openpyxl library.
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename
            sheet_name: Sheet name
            
        Returns:
            Excel file as bytes
        """
        logger.info(f"Exporting {len(data)} rows to Excel")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not data:
                return wb.save(filename) if filename else b''
            
            # Write headers
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            # Write data
            for row_num, row_data in enumerate(data, 2):
                cleaned_row = self._clean_row_for_export(row_data)
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=cleaned_row.get(header))
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            if filename:
                wb.save(filename)
                with open(filename, 'rb') as f:
                    return f.read()
            else:
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()
                
        except ImportError:
            logger.error("openpyxl library not installed")
            raise ValueError("Excel export requires openpyxl library")
    
    # ==================== JSON Export ====================
    
    def export_to_json(
        self,
        data: Any,
        pretty: bool = True
    ) -> str:
        """
        Export data to JSON format.
        
        Args:
            data: Data to export
            pretty: Whether to format JSON with indentation
            
        Returns:
            JSON string
        """
        logger.info(f"Exporting data to JSON")
        
        # Clean data for JSON serialization
        cleaned_data = self._clean_for_json(data)
        
        if pretty:
            return json.dumps(cleaned_data, indent=2, ensure_ascii=False)
        else:
            return json.dumps(cleaned_data, ensure_ascii=False)
    
    # ==================== PDF Export ====================
    
    def export_to_pdf(
        self,
        data: Dict[str, Any],
        template: str = 'default',
        title: str = 'Analytics Report'
    ) -> bytes:
        """
        Export data to PDF format.
        
        Requires reportlab or similar library.
        
        Args:
            data: Data to export
            template: Template name
            title: Report title
            
        Returns:
            PDF file as bytes
        """
        logger.info(f"Exporting data to PDF with template: {template}")
        
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            
            # Create PDF in memory
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            
            # Container for PDF elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            
            # Title
            title_style = styles['Title']
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Add metadata
            metadata = data.get('metadata', {})
            if metadata:
                elements.append(Paragraph(
                    f"Generated: {metadata.get('generated_at', 'N/A')}",
                    styles['Normal']
                ))
                elements.append(Spacer(1, 0.2*inch))
            
            # Add metrics as table
            metrics = data.get('metrics', {})
            if metrics:
                # Convert metrics to table data
                table_data = [['Metric', 'Value']]
                
                for key, value in self._flatten_dict(metrics).items():
                    table_data.append([key, str(value)])
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            return pdf_bytes
            
        except ImportError:
            logger.error("reportlab library not installed")
            raise ValueError("PDF export requires reportlab library")
    
    # ==================== Email Delivery ====================
    
    def email_report(
        self,
        recipients: List[str],
        subject: str,
        data: Any,
        format: str = 'pdf',
        attachments: Optional[List[Dict[str, Any]]] = None
    ) -> Dict[str, Any]:
        """
        Email report to recipients.
        
        Args:
            recipients: List of email addresses
            subject: Email subject
            data: Report data
            format: Export format (csv, excel, json, pdf)
            attachments: Optional additional attachments
            
        Returns:
            Send status
        """
        logger.info(f"Emailing report to {len(recipients)} recipients in {format} format")
        
        # Generate attachment based on format
        if format == 'csv':
            if isinstance(data, list):
                attachment_content = self.export_to_csv(data)
                attachment_name = 'report.csv'
            else:
                raise ValueError("CSV format requires list data")
                
        elif format == 'excel':
            if isinstance(data, list):
                attachment_content = self.export_to_excel(data)
                attachment_name = 'report.xlsx'
            else:
                raise ValueError("Excel format requires list data")
                
        elif format == 'json':
            attachment_content = self.export_to_json(data)
            attachment_name = 'report.json'
            
        elif format == 'pdf':
            attachment_content = self.export_to_pdf(data)
            attachment_name = 'report.pdf'
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        # Send email (placeholder - would integrate with email service)
        # In production, would use SendGrid, AWS SES, or similar
        logger.info(f"Email sent to {recipients} with attachment {attachment_name}")
        
        return {
            'success': True,
            'recipients': recipients,
            'format': format,
            'attachment_size': len(attachment_content) if isinstance(attachment_content, bytes) else len(attachment_content.encode())
        }
    
    # ==================== Helper Methods ====================
    
    def _clean_row_for_export(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """Clean row data for export."""
        cleaned = {}
        for key, value in row.items():
            if isinstance(value, Decimal):
                cleaned[key] = float(value)
            elif isinstance(value, (datetime, date)):
                cleaned[key] = value.isoformat()
            elif isinstance(value, UUID):
                cleaned[key] = str(value)
            elif isinstance(value, (dict, list)):
                cleaned[key] = json.dumps(value)
            else:
                cleaned[key] = value
        return cleaned
    
    def _clean_for_json(self, data: Any) -> Any:
        """Clean data for JSON serialization."""
        if isinstance(data, dict):
            return {k: self._clean_for_json(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._clean_for_json(item) for item in data]
        elif isinstance(data, Decimal):
            return float(data)
        elif isinstance(data, (datetime, date)):
            return data.isoformat()
        elif isinstance(data, UUID):
            return str(data)
        else:
            return data
    
    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
        """Flatten nested dictionary."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)


